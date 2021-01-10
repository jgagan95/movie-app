from openpyxl import load_workbook
import os, json
os.chdir('C:\\Users\\Josh Agan\\Documents\\movie-app')

workbook = load_workbook('movies.xlsx')
sheet = workbook['Sheet1']
movie_list = []

for row in sheet.iter_rows(min_row=1, max_col=5, max_row=356):
  movies = list(map(lambda num: num.value, row))
  genres = [genre for genre in movies[1:] if genre != None]
  movie_dictionaries = {'title': movies[0], 'genres': genres}
  movie_list.append(movie_dictionaries)

with open('movies.js', 'w') as output:
  output.write('let movies = %s' % json.dumps(movie_list))